from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import login, authenticate
from django.contrib import messages
from django.utils import timezone
from datetime import datetime, time, timedelta
from .models import ChecklistBase, SubgroupEntry, Verification, Shift, User
from .forms import ChecklistBaseForm, SubgroupEntryForm, SubgroupVerificationForm, VerificationForm, ConcernForm, UserRegistrationForm
from django.db.models import Count, Q
from django.utils import timezone
from datetime import timedelta
from django.utils import timezone
from django.db.models import Count, Q, Avg
from datetime import timedelta
import json
from django.shortcuts import render
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils import timezone
from datetime import timedelta
from itertools import chain
import json
from django.shortcuts import render



# Authentication Views
def register_user(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, 'Registration successful!')
            return redirect('dashboard')
        else:
            print(form.errors)  # Debugging: Prints errors in console
            messages.error(request, 'Registration failed. Please correct the errors below.')
    else:
        form = UserRegistrationForm()
    return render(request, 'main/register.html', {'form': form})

def user_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid username or password.')
    
    return render(request, 'main/login.html')

# Utility Functions
def get_current_shift(user):
    """Get or create current shift for user"""
    current_time = timezone.localtime()
    shift_type = 'day' if 8 <= current_time.hour < 20 else 'night'
    
    try:
        # Try to get existing shift
        shift = Shift.objects.get(
            date=current_time.date(),
            shift_type=shift_type
        )
    except Shift.DoesNotExist:
        # Get supervisors (or create dummy ones if needed)
        shift_supervisor = User.objects.filter(user_type='shift_supervisor').first()
        quality_supervisor = User.objects.filter(user_type='quality_supervisor').first()
        
        # If no supervisors exist, create default ones
        if not shift_supervisor:
            shift_supervisor = User.objects.create_user(
                username='default_shift_supervisor',
                password='defaultpass123',
                user_type='shift_supervisor'
            )
        
        if not quality_supervisor:
            quality_supervisor = User.objects.create_user(
                username='default_quality_supervisor',
                password='defaultpass123',
                user_type='quality_supervisor'
            )
        
        # Create new shift
        shift = Shift.objects.create(
            date=current_time.date(),
            shift_type=shift_type,
            operator=user,
            shift_supervisor=shift_supervisor,
            quality_supervisor=quality_supervisor
        )
    
    return shift

def check_time_gap(last_subgroup):
    """Check if enough time (2 hours) has passed since last subgroup"""
    if not last_subgroup:
        return True
    
    time_difference = timezone.now() - last_subgroup.timestamp
    return time_difference >= timedelta(hours=2)

# Main Views
@login_required
def dashboard(request):
    """Route to appropriate dashboard based on user type"""
    try:
        if request.user.user_type == 'operator':
            return operator_dashboard(request)
        elif request.user.user_type == 'shift_supervisor':
            return supervisor_dashboard(request)
        elif request.user.user_type == 'quality_supervisor':
            return quality_dashboard(request)
        else:
            # Default to operator dashboard if user_type is not set
            return operator_dashboard(request)
    except Exception as e:
        messages.error(request, f"Error loading dashboard: {str(e)}")
        return render(request, 'main/error.html', {'error': str(e)})

from django.contrib.auth.decorators import login_required
from django.utils import timezone
from datetime import timedelta

from datetime import datetime, timedelta
from django.utils import timezone

@login_required
def operator_dashboard(request):
    current_datetime = timezone.localtime(timezone.now())  # Use localtime
    current_date = current_datetime.date()
    current_time = current_datetime.time()
    
    # Shift determination - check exact hour
    hour = current_datetime.hour
    is_day_shift = (8 <= hour < 20)  # Between 8 AM and 8 PM
    current_shift = 'day' if is_day_shift else 'night'
    current_shift_display = 'Day Shift (8 AM - 8 PM)' if is_day_shift else 'Night Shift (8 PM - 8 AM)'
    
    # Debug time info (you can remove this later)
    print(f"Current time: {current_time}, Hour: {hour}, Is day shift: {is_day_shift}")
    
    # Get active checklist
    active_checklist = ChecklistBase.objects.filter(
        shift__operator=request.user,
        shift__date=current_date,
        shift__shift_type=current_shift,
        status='pending'
    ).prefetch_related('subgroup_entries').first()
    
    # Calculate next subgroup time and remaining time
    next_subgroup_time = None
    time_remaining = None
    if active_checklist:
        last_subgroup = active_checklist.subgroup_entries.order_by('-timestamp').first()
        if last_subgroup:
            # Use localtime for comparison
            next_time = timezone.localtime(last_subgroup.timestamp) + timedelta(hours=2)
            if next_time > current_datetime:
                time_remaining = next_time - current_datetime

    # Check if can add subgroup
    can_add_subgroup = False
    if active_checklist:
        subgroup_count = active_checklist.subgroup_entries.count()
        if subgroup_count < 6:
            if not last_subgroup:
                can_add_subgroup = True
            else:
                # Use localtime for comparison
                time_since_last = current_datetime - timezone.localtime(last_subgroup.timestamp)
                can_add_subgroup = time_since_last >= timedelta(hours=2)

    context = {
        'current_date': current_date,
        'current_time': current_time,
        'current_shift': current_shift_display,
        'active_checklist': active_checklist,
        'can_create_new': not active_checklist,
        'can_add_subgroup': can_add_subgroup,
        'next_subgroup_time': next_subgroup_time,
        'time_remaining': time_remaining
    }
    
    return render(request, 'main/operator_dashboard.html', context)

def check_time_gap(last_subgroup):
    """Helper function to check if enough time has passed since last subgroup"""
    if not last_subgroup:
        return True
    current_time = timezone.localtime(timezone.now())
    last_time = timezone.localtime(last_subgroup.timestamp)
    time_difference = current_time - last_time
    return time_difference >= timedelta(hours=2)



@login_required
@user_passes_test(lambda u: u.user_type == 'operator')
def create_checklist(request):
    """Create initial checklist with one-time entries"""
    current_shift = get_current_shift(request.user)
    
    # Check if checklist already exists for current shift
    if ChecklistBase.objects.filter(shift=current_shift).exists():
        messages.error(request, 'A checklist already exists for this shift')
        return redirect('operator_dashboard')
    
    if request.method == 'POST':
        form = ChecklistBaseForm(request.POST)
        if form.is_valid():
            try:
                checklist = form.save(commit=False)
                checklist.shift = current_shift
                checklist.save()
                messages.success(request, 'Checklist created successfully!')
                return redirect('add_subgroup', checklist_id=checklist.id)
            except Exception as e:
                messages.error(request, f'Error creating checklist: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = ChecklistBaseForm()
    
    return render(request, 'main/create_checklist.html', {
        'form': form,
        'current_shift': current_shift
    })

    
@login_required
@user_passes_test(lambda u: u.user_type == 'operator')
def add_subgroup(request, checklist_id):
    checklist = get_object_or_404(ChecklistBase, id=checklist_id)
    
    # Verify ownership and status
    if checklist.shift.operator != request.user:
        messages.error(request, 'You do not have permission to modify this checklist')
        return redirect('operator_dashboard')
    
    if checklist.status != 'pending':
        messages.error(request, 'Cannot add subgroups to a verified checklist')
        return redirect('checklist_detail', checklist_id=checklist.id)
    
    # Get latest subgroup 
    last_subgroup = checklist.subgroup_entries.order_by('-subgroup_number').first()
    current_subgroup = (last_subgroup.subgroup_number + 1 if last_subgroup else 1)
    
    if current_subgroup > 6:
        messages.error(request, 'Maximum number of subgroups (6) reached')
        return redirect('checklist_detail', checklist_id=checklist.id)
    
    # Check time gap if there's a previous subgroup
    if last_subgroup:
        time_since_last = timezone.now() - last_subgroup.timestamp
        if time_since_last < timedelta(hours=2):
            remaining_time = timedelta(hours=2) - time_since_last
            messages.warning(request, f'Note: Recommended to wait {remaining_time.seconds//60} minutes before adding next subgroup')
    
    if request.method == 'POST':
        form = SubgroupEntryForm(request.POST)
        if form.is_valid():
            subgroup = form.save(commit=False)
            subgroup.checklist = checklist
            subgroup.subgroup_number = current_subgroup
            subgroup.verification_status = 'pending'
            
            # Check for out-of-range values and add warnings
            warnings = []
            
            uv_vacuum = float(form.cleaned_data['uv_vacuum_test'])
            if not (-43 <= uv_vacuum <= -35):
                warnings.append(f'UV vacuum test value {uv_vacuum} kPa is outside recommended range (-43 to -35 kPa)')
            
            uv_flow = float(form.cleaned_data['uv_flow_value'])
            if not (30 <= uv_flow <= 40):
                warnings.append(f'UV flow value {uv_flow} LPM is outside recommended range (30-40 LPM)')
            
            # Save the subgroup regardless of warnings
            subgroup.save()
            
            # Show warnings as info messages
            for warning in warnings:
                messages.warning(request, warning)
            
            # Show message about pending verification
            if last_subgroup and last_subgroup.verification_status == 'pending':
                messages.info(request, f'Note: Subgroup {last_subgroup.subgroup_number} is still pending verification')
            
            messages.success(request, f'Subgroup {current_subgroup} added successfully')
            return redirect('checklist_detail', checklist_id=checklist.id)
    else:
        form = SubgroupEntryForm()
        # Show informational message about pending verification
        if last_subgroup and last_subgroup.verification_status == 'pending':
            messages.info(request, f'Note: Previous subgroup {last_subgroup.subgroup_number} is pending verification')
    
    return render(request, 'main/add_subgroup.html', {
        'form': form,
        'checklist': checklist,
        'current_subgroup': current_subgroup,
        'previous_subgroup_status': last_subgroup.verification_status if last_subgroup else None
    })    

@login_required
def validate_subgroup(request):
    """API endpoint for validating subgroup data"""
    if request.method == 'POST':
        data = request.POST
        warnings = []
        
        try:
            # Check UV vacuum test
            uv_vacuum = float(data.get('uv_vacuum_test', 0))
            if not (-43 <= uv_vacuum <= -35):
                warnings.append({
                    'field': 'uv_vacuum_test',
                    'message': f'UV vacuum test value {uv_vacuum} kPa is outside recommended range (-43 to -35 kPa)',
                    'value': uv_vacuum,
                    'recommended_range': {'min': -43, 'max': -35}
                })
            
            # Check UV flow value
            uv_flow = float(data.get('uv_flow_value', 0))
            if not (30 <= uv_flow <= 40):
                warnings.append({
                    'field': 'uv_flow_value',
                    'message': f'UV flow value {uv_flow} LPM is outside recommended range (30-40 LPM)',
                    'value': uv_flow,
                    'recommended_range': {'min': 30, 'max': 40}
                })
            
            return JsonResponse({
                'has_warnings': len(warnings) > 0,
                'warnings': warnings,
                'is_valid': True  # Always valid as we're only showing warnings
            })
            
        except (TypeError, ValueError) as e:
            return JsonResponse({
                'is_valid': False,
                'errors': ['Please enter valid numerical values']
            })
    
    return JsonResponse({'error': 'Invalid request method'}, status=400)

@login_required
def add_concern(request, checklist_id):
    checklist = get_object_or_404(ChecklistBase, id=checklist_id)
    
    # Only allow adding concerns if checklist is still pending
    if checklist.status not in ['pending', 'supervisor_approved']:
        messages.error(request, 'Cannot add concerns to a completed checklist')
        return redirect('checklist_detail', checklist_id=checklist.id)
    
    if request.method == 'POST':
        form = ConcernForm(request.POST)
        if form.is_valid():
            concern = form.save(commit=False)
            concern.checklist = checklist
            concern.save()
            messages.success(request, 'Concern added successfully')
            return redirect('checklist_detail', checklist_id=checklist.id)
    else:
        form = ConcernForm()
    
    return render(request, 'main/add_concern.html', {
        'form': form,
        'checklist': checklist
    })
    
@login_required
@user_passes_test(lambda u: u.user_type == 'operator')
def edit_subgroup(request, checklist_id, subgroup_id):
    checklist = get_object_or_404(ChecklistBase, id=checklist_id)
    subgroup = get_object_or_404(SubgroupEntry, id=subgroup_id, checklist=checklist)
    
    # Verify ownership and status
    if checklist.shift.operator != request.user:
        messages.error(request, 'You do not have permission to modify this subgroup')
        return redirect('checklist_detail', checklist_id=checklist.id)
    
    if checklist.status != 'pending':
        messages.error(request, 'Cannot edit subgroups of a verified checklist')
        return redirect('checklist_detail', checklist_id=checklist.id)
    
    if request.method == 'POST':
        form = SubgroupEntryForm(request.POST, instance=subgroup)
        if form.is_valid():
            form.save()
            messages.success(request, 'Subgroup updated successfully')
            return redirect('checklist_detail', checklist_id=checklist.id)
    else:
        form = SubgroupEntryForm(instance=subgroup)
    
    return render(request, 'main/edit_subgroup.html', {
        'form': form,
        'checklist': checklist,
        'subgroup': subgroup
    })
        
@login_required
@user_passes_test(lambda u: u.user_type == 'shift_supervisor')
def supervisor_dashboard(request):
    current_datetime = timezone.now()
    current_date = current_datetime.date()
    current_time = current_datetime.time()
    is_day_shift = (8 <= current_time.hour < 20)
    current_shift = 'day' if is_day_shift else 'night'
    current_shift_display = 'Day Shift (8 AM - 8 PM)' if is_day_shift else 'Night Shift (8 PM - 8 AM)'

    # Get subgroups that need supervisor verification (pending)
    pending_subgroups = SubgroupEntry.objects.filter(
        verification_status='pending',
        checklist__shift__shift_supervisor=request.user,
        checklist__shift__date=current_date
    ).select_related(
        'checklist__shift__operator'
    ).prefetch_related(
        'verifications'
    ).order_by('-timestamp')

    # Get supervisor's verifications from today
    todays_verifications = SubgroupVerification.objects.filter(
        verified_by=request.user,
        verifier_type='supervisor',
        verified_at__date=current_date
    ).select_related(
        'subgroup__checklist__shift__operator'
    ).order_by('-verified_at')

    context = {
        'current_date': current_date,
        'current_time': current_time,
        'current_shift': current_shift_display,
        'pending_verifications': pending_subgroups,
        'verified_entries': todays_verifications,
        'verification_summary': {
            'pending_count': pending_subgroups.count(),
            'verified_today': todays_verifications.count()
        }
    }
    
    return render(request, 'main/supervisor_dashboard.html', context)





def calculate_average_verification_time(entries):
    """Calculate average time taken for verification"""
    verified_entries = entries.filter(status='supervisor_approved')
    if not verified_entries:
        return 0
        
    total_time = timedelta()
    count = 0
    
    for entry in verified_entries:
        if entry.created_at and entry.supervisor_verified_at:
            time_diff = entry.supervisor_verified_at - entry.created_at
            total_time += time_diff
            count += 1
            
    if count == 0:
        return 0
        
    average_seconds = total_time.total_seconds() / count
    return round(average_seconds / 60)  # Return in minutes

def calculate_completion_rate(entries):
    """Calculate the rate of checklists that have all 6 subgroups"""
    total = entries.count()
    if total == 0:
        return 0
        
    completed = entries.annotate(
        subgroup_count=Count('subgroup_entries')
    ).filter(subgroup_count=6).count()
    
    return round((completed / total) * 100)

@login_required
@user_passes_test(lambda u: u.user_type == 'shift_supervisor')
def verify_checklist(request, checklist_id):
    checklist = get_object_or_404(ChecklistBase, id=checklist_id)
    
    if request.method == 'POST':
        form = VerificationForm(request.POST)
        if form.is_valid():
            verification = form.save(commit=False)
            verification.checklist = checklist
            verification.supervisor = request.user
            verification.save()
            
            checklist.status = 'supervisor_approved'
            checklist.save()
            
            messages.success(request, 'Checklist verified successfully')
            return redirect('supervisor_dashboard')
    else:
        form = VerificationForm()
    
    return render(request, 'main/verify_checklist.html', {
        'form': form,
        'checklist': checklist
    })


def calculate_model_stats(entries):
    """Calculate statistics for each model"""
    model_stats = []
    for model_code, model_name in ChecklistBase.MODEL_CHOICES:
        model_entries = entries.filter(selected_model=model_code)
        model_total = model_entries.count()
        
        if model_total > 0:
            stats = {
                'model': model_code,
                'total': model_total,
                'approved': model_entries.filter(status='quality_approved').count(),
                'rejected': model_entries.filter(status='rejected').count()
            }
            stats['rate'] = round((stats['approved'] / model_total) * 100)
            model_stats.append(stats)
    
    return model_stats

def calculate_trend_data(current_date):
    """Calculate 7-day trend data"""
    trend_dates = []
    trend_rates = []
    
    for i in range(7, -1, -1):
        date = current_date - timedelta(days=i)
        entries = ChecklistBase.objects.filter(
            shift__date=date,
            status__in=['quality_approved', 'rejected']
        )
        total = entries.count()
        if total > 0:
            approved = entries.filter(status='quality_approved').count()
            rate = round((approved / total) * 100)
        else:
            rate = 0
        trend_dates.append(date.strftime('%Y-%m-%d'))
        trend_rates.append(rate)
    
    return {
        'dates': trend_dates,
        'rates': trend_rates
    }

def calculate_average_verification_time(verifications):
    """Calculate average time taken for quality verification"""
    total_time = timedelta()
    count = 0
    
    for verification in verifications:
        if hasattr(verification, 'verifications'):
            for v in verification.verifications.all():
                if hasattr(v, 'supervisor_verified_at') and v.quality_supervisor_verified_at:
                    time_diff = v.quality_supervisor_verified_at - v.supervisor_verified_at
                    total_time += time_diff
                    count += 1
            
    if count == 0:
        return 0
        
    return round(total_time.total_seconds() / count / 60)  # Return in minutes

def calculate_quality_stats(entries):
    """Calculate quality statistics"""
    stats = {
        'approved': entries.filter(status='quality_approved').count(),
        'rejected': entries.filter(status='rejected').count(),
        'pending': entries.filter(status__in=['pending', 'supervisor_approved']).count()
    }
    
    total_processed = stats['approved'] + stats['rejected']
    stats['approval_rate'] = round((stats['approved'] / total_processed * 100) if total_processed > 0 else 0)
    
    return stats

def process_measurements(checklists):
    """Process and validate measurements for a list of checklists"""
    processed_entries = []
    
    for checklist in checklists:
        entry_data = {
            'id': checklist.id,
            'created_at': checklist.created_at,
            'shift': checklist.shift,
            'selected_model': checklist.selected_model,
            'subgroup_count': checklist.subgroup_entries.count(),
            'all_measurements_ok': True,
            'measurement_issues': [],
            'critical_issues': []
        }

        # Process subgroups
        for subgroup in checklist.subgroup_entries.all():
            # Validate UV Vacuum Test
            if not (-43 <= subgroup.uv_vacuum_test <= -35):
                entry_data['measurement_issues'].append(
                    f"Subgroup {subgroup.subgroup_number}: UV vacuum test out of range ({subgroup.uv_vacuum_test})"
                )
                entry_data['all_measurements_ok'] = False

            # Validate UV Flow Value
            if not (30 <= subgroup.uv_flow_value <= 40):
                entry_data['measurement_issues'].append(
                    f"Subgroup {subgroup.subgroup_number}: UV flow value out of range ({subgroup.uv_flow_value})"
                )
                entry_data['all_measurements_ok'] = False

            # Check other critical values
            if not ((4.5 <= checklist.line_pressure <= 5.5) and 
                   (11 <= checklist.uv_flow_input_pressure <= 15) and
                   (0.25 <= checklist.test_pressure_vacuum <= 0.3)):
                entry_data['critical_issues'].append({
                    'severity': 'critical',
                    'message': 'Critical measurements out of range',
                    'measurements': {
                        'line_pressure': checklist.line_pressure,
                        'uv_flow_input_pressure': checklist.uv_flow_input_pressure,
                        'test_pressure_vacuum': checklist.test_pressure_vacuum
                    }
                })

        processed_entries.append(entry_data)

    return processed_entries

def validate_base_measurements(entry):
    """Validate base measurements"""
    measurements = {
        'line_pressure_ok': False,
        'uv_flow_input_ok': False,
        'test_pressure_ok': False
    }
    
    try:
        # Line Pressure validation
        if hasattr(entry, 'line_pressure'):
            line_pressure = float(entry.line_pressure)
            measurements['line_pressure_ok'] = 4.5 <= line_pressure <= 5.5
            if not measurements['line_pressure_ok']:
                entry.critical_issues.append(f"Line pressure critical: {line_pressure}")
        
        # UV Flow Input validation
        if hasattr(entry, 'uv_flow_input_pressure'):
            uv_pressure = float(entry.uv_flow_input_pressure)
            measurements['uv_flow_input_ok'] = 11 <= uv_pressure <= 15
            if not measurements['uv_flow_input_ok']:
                entry.critical_issues.append(f"UV flow input pressure critical: {uv_pressure}")
        
        # Test Pressure validation
        if hasattr(entry, 'test_pressure_vacuum'):
            test_pressure = float(entry.test_pressure_vacuum)
            measurements['test_pressure_ok'] = 0.25 <= test_pressure <= 0.3
            if not measurements['test_pressure_ok']:
                entry.critical_issues.append(f"Test pressure critical: {test_pressure}")
    except (ValueError, TypeError):
        entry.critical_issues.append("Invalid measurement values")
    
    return measurements

def validate_subgroup_measurements(entry):
    """Validate subgroup measurements"""
    subgroup_validations = []
    
    try:
        for subgroup in entry.subgroup_entries.all():
            validation = {
                'subgroup_number': subgroup.subgroup_number,
                'uv_vacuum_test_ok': False,
                'uv_flow_value_ok': False,
                'all_ok': True
            }
            
            # UV Vacuum Test validation
            try:
                uv_vacuum = float(subgroup.uv_vacuum_test)
                validation['uv_vacuum_test_ok'] = -43 <= uv_vacuum <= -35
                if not validation['uv_vacuum_test_ok']:
                    entry.measurement_issues.append(
                        f"Subgroup {subgroup.subgroup_number}: UV vacuum test out of range ({uv_vacuum})"
                    )
            except (ValueError, TypeError):
                validation['all_ok'] = False
            
            # UV Flow Value validation
            try:
                uv_flow = float(subgroup.uv_flow_value)
                validation['uv_flow_value_ok'] = 30 <= uv_flow <= 40
                if not validation['uv_flow_value_ok']:
                    entry.measurement_issues.append(
                        f"Subgroup {subgroup.subgroup_number}: UV flow value out of range ({uv_flow})"
                    )
            except (ValueError, TypeError):
                validation['all_ok'] = False
            
            subgroup_validations.append(validation)
    except Exception as e:
        entry.measurement_issues.append(f"Error processing subgroups: {str(e)}")
    
    return subgroup_validations

def get_critical_issues(in_progress, pending):
    """Get critical issues from both in-progress and pending checklists"""
    critical_issues = []
    
    for entries in [in_progress, pending]:
        for entry in entries:
            if entry['critical_issues']:
                critical_issues.append({
                    'checklist_id': entry['id'],
                    'operator': entry['original_entry'].shift.operator.username,
                    'model': entry['original_entry'].selected_model,
                    'issues': entry['critical_issues']
                })
    

@login_required
@user_passes_test(lambda u: u.user_type == 'quality_supervisor')
def quality_dashboard(request):
    current_datetime = timezone.now()
    current_date = current_datetime.date()
    current_time = current_datetime.time()
    is_day_shift = (8 <= current_time.hour < 20)
    current_shift = 'day' if is_day_shift else 'night'
    current_shift_display = 'Day Shift (8 AM - 8 PM)' if is_day_shift else 'Night Shift (8 PM - 8 AM)'

    # Get subgroups that need quality verification (where supervisor has verified)
    pending_subgroups = SubgroupEntry.objects.filter(
        verification_status='supervisor_verified',
        checklist__shift__quality_supervisor=request.user,
        checklist__shift__date=current_date
    ).select_related(
        'checklist__shift__operator',
        'checklist__shift__shift_supervisor'
    ).prefetch_related(
        'verifications'
    ).order_by('-timestamp')

    # Get quality verifications done today
    todays_verifications = SubgroupVerification.objects.filter(
        verified_by=request.user,
        verifier_type='quality',
        verified_at__date=current_date
    ).select_related(
        'subgroup__checklist__shift__operator',
        'subgroup__checklist__shift__shift_supervisor'
    ).order_by('-verified_at')

    context = {
        'current_date': current_date,
        'current_time': current_time,
        'current_shift': current_shift_display,
        'pending_verifications': pending_subgroups,
        'verified_entries': todays_verifications,
        'verification_stats': {
            'pending_count': pending_subgroups.count(),
            'approved_today': todays_verifications.filter(status='approved').count(),
            'rejected_today': todays_verifications.filter(status='rejected').count()
        }
    }
    
    return render(request, 'main/quality_dashboard.html', context)


@login_required
def reports_dashboard(request):
    current_date = timezone.now().date()
    start_of_week = current_date - timedelta(days=current_date.weekday())
    start_of_month = current_date.replace(day=1)

    # Get statistics
    daily_stats = ChecklistBase.objects.filter(
        shift__date=current_date
    ).aggregate(
        total=Count('id'),
        approved=Count('id', filter=Q(status='quality_approved')),
        rejected=Count('id', filter=Q(status='rejected')),
        pending=Count('id', filter=Q(status='pending'))
    )

    weekly_stats = ChecklistBase.objects.filter(
        shift__date__gte=start_of_week,
        shift__date__lte=current_date
    ).aggregate(
        total=Count('id'),
        approved=Count('id', filter=Q(status='quality_approved')),
        rejected=Count('id', filter=Q(status='rejected'))
    )

    monthly_stats = ChecklistBase.objects.filter(
        shift__date__gte=start_of_month,
        shift__date__lte=current_date
    ).aggregate(
        total=Count('id'),
        approved=Count('id', filter=Q(status='quality_approved')),
        rejected=Count('id', filter=Q(status='rejected'))
    )

    # Get model-wise statistics
    model_stats = ChecklistBase.objects.values('selected_model').annotate(
        total=Count('id'),
        approved=Count('id', filter=Q(status='quality_approved')),
        rejected=Count('id', filter=Q(status='rejected'))
    )

    # Calculate approval rates
    for stats in [daily_stats, weekly_stats, monthly_stats]:
        total_verified = stats['approved'] + stats['rejected']
        stats['approval_rate'] = (
            round((stats['approved'] / total_verified) * 100)
            if total_verified > 0 else 0
        )

    # Get recent activity
    recent_activity = ChecklistBase.objects.filter(
        status__in=['quality_approved', 'rejected']
    ).order_by('-created_at')[:10]

    context = {
        'daily_stats': daily_stats,
        'weekly_stats': weekly_stats,
        'monthly_stats': monthly_stats,
        'model_stats': model_stats,
        'recent_activity': recent_activity,
        'current_date': current_date
    }

    return render(request, 'main/reports/reports_dashboard.html', context)


from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def export_checklist_excel(request, checklist_id):
    checklist = get_object_or_404(
        ChecklistBase.objects.select_related(
            'shift__operator',
            'shift__shift_supervisor',
            'shift__quality_supervisor'
        ).prefetch_related(
            'subgroup_entries__verifications',
        ), 
        id=checklist_id
    )
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Quality Control Data"
    
    # Styles
    header_style = {
        'fill': PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid'),
        'font': Font(bold=True, color='FFFFFF'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }
    
    subheader_style = {
        'fill': PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid'),
        'font': Font(bold=True),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }
    
    pass_style = {
        'fill': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'font': Font(color='006100'),
        'alignment': Alignment(horizontal='center', vertical='center')
    }
    
    fail_style = {
        'fill': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        'font': Font(color='9C0006'),
        'alignment': Alignment(horizontal='center', vertical='center')
    }
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def apply_style(cell, style):
        cell.fill = style['fill']
        cell.font = style['font']
        cell.alignment = style['alignment']
        cell.border = border

    # Header row for date and shift info
    ws['A1'] = 'Date:'
    ws['B1'] = checklist.shift.date.strftime('%Y-%m-%d')
    ws['D1'] = 'Shift:'
    ws['E1'] = checklist.shift.get_shift_type_display()

    # Subgroup headers
    ws['A3'] = 'CHARACTERISTICS'
    apply_style(ws['A3'], header_style)
    
    # Get ordered subgroups
    subgroups = checklist.subgroup_entries.all().order_by('subgroup_number')
    
    # Create subgroup headers
    for sg_num in range(1, 5):
        start_col = 2 + (sg_num - 1) * 5
        for i in range(5):
            col = get_column_letter(start_col + i)
            ws.merge_cells(f'{col}3:{col}4')
            ws[f'{col}3'] = f'Sub Group {sg_num}\nSample {i+1}'
            apply_style(ws[f'{col}3'], header_style)

    # Record Time row
    current_row = 4
    ws['A4'] = 'Record Time'
    for sg in subgroups:
        col_start = 2 + (sg.subgroup_number - 1) * 5
        time_str = sg.timestamp.strftime('%H:%M')
        for i in range(5):
            col = get_column_letter(col_start + i)
            ws[f'{col}5'] = time_str

    current_row = 5

    # Single value entries
    single_entries = [
        ('Program selection on HMI (HMI से Program select करना है)', checklist.selected_model),
        ('Line pressure (4.5 - 5.5 bar)', checklist.line_pressure),
        ('O-ring conditon(UV Flow check sealing area), should not be damaged', checklist.oring_condition),
        ('UV Flow input Test Pressure(13+/- 2 KPa) (11-15) kPA', checklist.uv_flow_input_pressure),
    ]

    for entry_name, value in single_entries:
        ws[f'A{current_row}'] = entry_name
        ws[f'B{current_row}'] = value
        current_row += 1

    # Measurements with 5 samples per subgroup
    measurements = [
        ('UV Vaccum Test range(-35 to -43 KPa)', 'uv_vacuum_test', lambda x: -43 <= x <= -35),
        ('UV Flow Value (30~40 LPM)(HMI)', 'uv_flow_value', lambda x: 30 <= x <= 40),
    ]

    for measurement_name, attr, validation_func in measurements:
        ws[f'A{current_row}'] = measurement_name
        for sg in subgroups:
            col_start = 2 + (sg.subgroup_number - 1) * 5
            value = getattr(sg, attr)
            for i in range(5):
                col = get_column_letter(col_start + i)
                ws[f'{col}{current_row}'] = value
                apply_style(ws[f'{col}{current_row}'], 
                          pass_style if validation_func(value) else fail_style)
        current_row += 1

    # Single-value fields without repetition
    for field_name, attr in [
        ('Master Verification for LVDT (OK/NG)', 'master_verification_lvdt'),
        ('Good and Bad master verification (refer EPVS)', 'good_bad_master_verification'),
        ('Test Pressure for Vacumm generation (0.25 ~ 0.3 Mpa)', 'test_pressure_vacuum'),
        ('Tool Alignmnet (Top & Bottom)', 'tool_alignment'),
        ('Tool Id: Top Tool ID : FMA-03-35- T05 (P703/U704/SA/FD/Gnome)', 'top_tool_id'),
        ('Bottom Tool ID: FMA-03-35-T06 (P703/U704/SA/FD)', 'bottom_tool_id'),
        ('UV Assy Stage 1 ID: FMA-03-35- T07 (P703/U704/SA/FD)', 'uv_assy_stage_id'),
        ('Retainer Part no - 42001878 (P703/U704/SA/FD)', 'retainer_part_no'),
        ('UV Clip Part No -  42000829 (P703/U704/SA/FD)', 'uv_clip_part_no'),
        ('Umbrella Part No - 25094588 (P703/U704/SA/FD/Gnome)', 'umbrella_part_no'),
        ('Retainer ID lubrication', 'retainer_id_lubrication'),
    ]:
        ws[f'A{current_row}'] = field_name
        value = getattr(checklist, attr)
        ws[f'B{current_row}'] = value
        if attr in ['master_verification_lvdt', 'good_bad_master_verification', 'tool_alignment', 'retainer_id_lubrication']:
            apply_style(ws[f'B{current_row}'], 
                       pass_style if value == 'OK' else fail_style)
        current_row += 1

    # Repeated checks with 5 samples per subgroup
    repeated_checks = [
        ('Umbrella Valve Assembly in Retainer in UV Assy Station', 'umbrella_valve_assembly'),
        ('UV Clip pressing -proper locking of 2 nos snap', 'uv_clip_pressing'),
        ('All workstations are clean (Y/N) वर्कस्टेशन साफ होना चाहिए (हाँ/ना)', 'workstation_clean'),
        ('Station Operator will confirm that every bin feeded on line is free from contamination (Y/N) PTGW_5.3_PC_GUR_03', 'bin_contamination_check')
    ]

    for check_name, attr in repeated_checks:
        ws[f'A{current_row}'] = check_name
        for sg in subgroups:
            col_start = 2 + (sg.subgroup_number - 1) * 5
            value = getattr(sg, attr)
            for i in range(5):
                col = get_column_letter(col_start + i)
                ws[f'{col}{current_row}'] = value
                apply_style(ws[f'{col}{current_row}'], 
                          pass_style if value in ['OK', 'Yes'] else fail_style)
        current_row += 1

    # Signatures section - one column per subgroup
    current_row += 1
    for signature_row, role in [
        ('Team Leader/Operator initial', 'operator'),
        ('Shift Supervisor Initials', 'shift_supervisor'),
        ('Quality Supervisor Initials', 'quality_supervisor')
    ]:
        ws[f'A{current_row}'] = signature_row
        for sg in subgroups:
            col_start = 2 + (sg.subgroup_number - 1) * 5
            name = getattr(checklist.shift, role).get_full_name() or getattr(checklist.shift, role).username
            for i in range(5):
                col = get_column_letter(col_start + i)
                ws[f'{col}{current_row}'] = name
        current_row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 60
    for i in range(1, 21):
        col = get_column_letter(i + 1)
        ws.column_dimensions[col].width = 12

    # Generate response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=quality_control_{checklist.id}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    wb.save(response)
    return response
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from datetime import timedelta
from django.db.models import Avg, Count
from .models import ChecklistBase, SubgroupEntry, Verification, Concern

@login_required
def checklist_detail(request, checklist_id):
    # Get checklist with related data
    checklist = get_object_or_404(
        ChecklistBase.objects.select_related(
            'shift__operator',
            'shift__shift_supervisor',
            'shift__quality_supervisor'
        ).prefetch_related(
            'subgroup_entries__verifications',
            'verifications',
            'concern_set'
        ), 
        id=checklist_id
    )
    
    # Get subgroups with validation status
    subgroups = process_subgroups(checklist.subgroup_entries.all().order_by('subgroup_number'))
    
    # Calculate subgroup metrics
    subgroup_metrics = calculate_subgroup_metrics(subgroups)
    
    # Check if new subgroup can be added
    can_add_subgroup = check_subgroup_addition(checklist, subgroups)
    
    # Get verification status and permissions
    verification_status = get_verification_status(request.user, checklist, subgroups)
    
    # Process measurement validations
    measurement_validation = validate_measurements(checklist, subgroups)
    
    # Get concerns with metadata
    concerns = process_concerns(checklist.concern_set.all())
    
    # Calculate timing metrics
    timing_metrics = calculate_timing_metrics(checklist, subgroups)
    
    context = {
        'checklist': checklist,
        'subgroups': subgroups,
        'can_add_subgroup': can_add_subgroup,
        'total_subgroups': len(subgroups),
        'remaining_subgroups': 6 - len(subgroups),
        'subgroup_metrics': subgroup_metrics,
        'verification_status': verification_status,
        'measurement_validation': measurement_validation,
        'concerns': concerns,
        'timing_metrics': timing_metrics,
        'user_permissions': get_user_permissions(request.user, checklist)
    }
    
    return render(request, 'main/checklist_detail.html', context)

def process_subgroups(subgroups):
    """Process and validate subgroups"""
    processed_subgroups = []
    for subgroup in subgroups:
        subgroup.validation_status = {
            'uv_vacuum_test_ok': -43 <= subgroup.uv_vacuum_test <= -35,
            'uv_flow_value_ok': 30 <= subgroup.uv_flow_value <= 40,
            'assembly_ok': subgroup.umbrella_valve_assembly == 'OK',
            'pressing_ok': subgroup.uv_clip_pressing == 'OK',
            'cleanliness_ok': subgroup.workstation_clean == 'Yes',
            'contamination_ok': subgroup.bin_contamination_check == 'Yes'
        }
        subgroup.all_checks_passed = all(subgroup.validation_status.values())
        
        # Add verification information
        subgroup.supervisor_verification = SubgroupVerification.objects.filter(
            subgroup=subgroup,
            verifier_type='supervisor'
        ).first()
        
        subgroup.quality_verification = SubgroupVerification.objects.filter(
            subgroup=subgroup,
            verifier_type='quality'
        ).first()
        
        processed_subgroups.append(subgroup)
    return processed_subgroups

def calculate_subgroup_metrics(subgroups):
    """Calculate metrics for subgroups"""
    return {
        'total_count': len(subgroups),
        'passed_count': sum(1 for s in subgroups if s.all_checks_passed),
        'vacuum_test_issues': sum(1 for s in subgroups if not s.validation_status['uv_vacuum_test_ok']),
        'flow_value_issues': sum(1 for s in subgroups if not s.validation_status['uv_flow_value_ok']),
        'completion_percentage': (len(subgroups) / 6) * 100 if subgroups else 0
    }

def check_subgroup_addition(checklist, subgroups):
    """Check if new subgroup can be added"""
    if checklist.status != 'pending':
        return False
        
    current_count = len(subgroups)
    if current_count >= 6:
        return False
        
    if not subgroups:
        return True
        
    last_subgroup = subgroups[-1]
    time_since_last = timezone.now() - last_subgroup.timestamp
    return time_since_last >= timedelta(hours=2)

def get_verification_status(user, checklist, subgroups):
    """Get verification status and permissions"""
    return {
        'can_verify_supervisor': (
            user.user_type == 'shift_supervisor' and 
            checklist.status == 'pending' and 
            len(subgroups) == 6
        ),
        'can_verify_quality': (
            user.user_type == 'quality_supervisor' and 
            checklist.status == 'supervisor_approved'
        ),
        'last_verification': checklist.verifications.last(),
        'verification_count': checklist.verifications.count()
    }

def validate_measurements(checklist, subgroups):
    """Validate all measurements"""
    base_measurements = {
        'line_pressure_ok': 4.5 <= checklist.line_pressure <= 5.5,
        'uv_flow_input_ok': 11 <= checklist.uv_flow_input_pressure <= 15,
        'test_pressure_ok': 0.25 <= checklist.test_pressure_vacuum <= 0.3
    }
    
    return {
        'base_measurements': base_measurements,
        'all_base_ok': all(base_measurements.values()),
        'subgroup_measurements_ok': all(s.all_checks_passed for s in subgroups),
        'critical_issues': [
            issue for issue in get_critical_issues(checklist, subgroups)
            if issue['severity'] == 'critical'
        ]
    }

def process_concerns(concerns):
    """Process concerns with additional metadata"""
    return [{
        'concern': concern,
        'is_resolved': bool(concern.action_taken),
        'requires_attention': not concern.action_taken
    } for concern in concerns]
    



def calculate_timing_metrics(checklist, subgroups):
    """Calculate timing related metrics"""
    if not subgroups:
        return None
        
    return {
        'total_duration': (subgroups[-1].timestamp - checklist.created_at).total_seconds() / 3600,
        'average_interval': calculate_average_interval(subgroups),
        'completion_rate': len(subgroups) / ((timezone.now() - checklist.created_at).total_seconds() / 3600)
    }

def get_user_permissions(user, checklist):
    """Get user-specific permissions"""
    return {
        'can_edit': user.user_type == 'operator' and checklist.status == 'pending',
        'can_verify': user.user_type in ['shift_supervisor', 'quality_supervisor'],
        'can_add_concern': checklist.status != 'quality_approved',
        'can_view_all': user.user_type in ['shift_supervisor', 'quality_supervisor']
    }

def calculate_average_interval(subgroups):
    """Calculate average time interval between subgroups"""
    if len(subgroups) < 2:
        return None
        
    intervals = []
    for i in range(1, len(subgroups)):
        interval = (subgroups[i].timestamp - subgroups[i-1].timestamp).total_seconds() / 3600
        intervals.append(interval)
    
    return sum(intervals) / len(intervals)

def get_critical_issues(checklist, subgroups):
    """Identify critical issues in measurements"""
    issues = []
    
    # Check base measurements
    if not (4.5 <= checklist.line_pressure <= 5.5):
        issues.append({
            'type': 'base_measurement',
            'measurement': 'line_pressure',
            'value': checklist.line_pressure,
            'severity': 'critical'
        })
    
    # Check subgroup measurements
    for subgroup in subgroups:
        if not (-43 <= subgroup.uv_vacuum_test <= -35):
            issues.append({
                'type': 'subgroup',
                'subgroup': subgroup.subgroup_number,
                'measurement': 'uv_vacuum_test',
                'value': subgroup.uv_vacuum_test,
                'severity': 'critical'
            })
    
    return issues
    
    
    
    
    
from .forms import SubgroupVerificationForm
from django.db import IntegrityError
from .models import (
    ChecklistBase, 
    SubgroupEntry, 
    SubgroupVerification, 
    Shift, 
    User
)

@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'])
def verify_subgroup_measurement(request, subgroup_id):
    subgroup = get_object_or_404(SubgroupEntry.objects.select_related('checklist'), id=subgroup_id)
    verifier_type = 'supervisor' if request.user.user_type == 'shift_supervisor' else 'quality'
    
    # Get existing verification
    existing_verification = SubgroupVerification.objects.filter(
        subgroup=subgroup,
        verifier_type=verifier_type
    ).first()
    
    # Check verification rules
    if verifier_type == 'quality':
        supervisor_verification = SubgroupVerification.objects.filter(
            subgroup=subgroup,
            verifier_type='supervisor'
        ).first()
        
        if not supervisor_verification or supervisor_verification.status == 'rejected':
            messages.error(request, 'Subgroup must be verified by supervisor first')
            return redirect('checklist_detail', checklist_id=subgroup.checklist.id)
    
    if request.method == 'POST':
        form = SubgroupVerificationForm(request.POST, instance=existing_verification)
        if form.is_valid():
            try:
                verification = form.save(commit=False)
                verification.subgroup = subgroup
                verification.verified_by = request.user
                verification.verifier_type = verifier_type
                verification.save()
                
                # Update subgroup status
                new_status = None
                if verification.status == 'rejected':
                    new_status = 'rejected'
                elif verifier_type == 'supervisor':
                    new_status = 'supervisor_verified'
                else:  # quality supervisor
                    new_status = 'quality_verified'
                    
                subgroup.verification_status = new_status
                subgroup.save()
                
                messages.success(request, 'Verification completed successfully')
                return redirect('checklist_detail', checklist_id=subgroup.checklist.id)
                
            except IntegrityError:
                messages.error(request, 'An error occurred during verification')
        else:
            messages.error(request, 'Please correct the errors below')
    else:
        form = SubgroupVerificationForm(instance=existing_verification)
    
    # Prepare measurements validation for template
    measurements = {
        'uv_vacuum_test': {
            'value': subgroup.uv_vacuum_test,
            'is_valid': -43 <= subgroup.uv_vacuum_test <= -35,
            'range': '-43 to -35 kPa'
        },
        'uv_flow_value': {
            'value': subgroup.uv_flow_value,
            'is_valid': 30 <= subgroup.uv_flow_value <= 40,
            'range': '30-40 LPM'
        },
        'assembly_ok': subgroup.umbrella_valve_assembly == 'OK',
        'pressing_ok': subgroup.uv_clip_pressing == 'OK',
        'cleanliness': {
            'workstation': subgroup.workstation_clean == 'Yes',
            'contamination': subgroup.bin_contamination_check == 'Yes'
        }
    }
    
    context = {
        'form': form,
        'subgroup': subgroup,
        'measurements': measurements,
        'verifier_type': verifier_type.title(),
        'existing_verification': existing_verification,
    }
    
    return render(request, 'main/verify_subgroup_measurement.html', context)    
    
    
@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'])
def edit_verification(request, verification_id):
    verification = get_object_or_404(SubgroupVerification, id=verification_id)
    
    # Check if user has permission to edit this verification
    if (request.user.user_type == 'shift_supervisor' and verification.verifier_type != 'supervisor') or \
       (request.user.user_type == 'quality_supervisor' and verification.verifier_type != 'quality'):
        messages.error(request, 'You do not have permission to edit this verification')
        return redirect('checklist_detail', checklist_id=verification.subgroup.checklist.id)
    
    if request.method == 'POST':
        form = SubgroupVerificationForm(request.POST, instance=verification)
        if form.is_valid():
            verification = form.save(commit=False)
            verification.verified_at = timezone.now()  # Update verification time
            verification.save()
            
            # Update subgroup status
            subgroup = verification.subgroup
            if verification.status == 'rejected':
                subgroup.verification_status = 'rejected'
            elif verification.verifier_type == 'supervisor':
                subgroup.verification_status = 'supervisor_verified'
            else:  # quality supervisor
                subgroup.verification_status = 'quality_verified'
            subgroup.save()
            
            messages.success(request, 'Verification updated successfully')
            return redirect('checklist_detail', checklist_id=verification.subgroup.checklist.id)
    else:
        form = SubgroupVerificationForm(instance=verification)
    
    context = {
        'form': form,
        'verification': verification,
        'subgroup': verification.subgroup,
        'measurements': get_subgroup_measurements(verification.subgroup),
    }
    
    return render(request, 'main/edit_verification.html', context)

def get_subgroup_measurements(subgroup):
    """Helper function to get formatted measurements for a subgroup"""
    return {
        'uv_vacuum_test': {
            'value': subgroup.uv_vacuum_test,
            'is_valid': -43 <= subgroup.uv_vacuum_test <= -35,
            'range': '-43 to -35 kPa'
        },
        'uv_flow_value': {
            'value': subgroup.uv_flow_value,
            'is_valid': 30 <= subgroup.uv_flow_value <= 40,
            'range': '30-40 LPM'
        },
        'assembly_ok': subgroup.umbrella_valve_assembly == 'OK',
        'pressing_ok': subgroup.uv_clip_pressing == 'OK',
        'cleanliness': {
            'workstation': subgroup.workstation_clean == 'Yes',
            'contamination': subgroup.bin_contamination_check == 'Yes'
        }
    }    
# Verification Views
@login_required
@user_passes_test(lambda u: u.user_type == 'shift_supervisor')
def supervisor_verify(request, checklist_id):
    checklist = get_object_or_404(ChecklistBase, id=checklist_id)
    
    # Verify if checklist can be verified
    if checklist.status != 'pending':
        messages.error(request, 'This checklist has already been verified')
        return redirect('supervisor_dashboard')
        
    if checklist.subgroup_entries.count() < 6:
        messages.error(request, 'Cannot verify incomplete checklist. All 6 subgroups must be completed.')
        return redirect('checklist_detail', checklist_id=checklist.id)
    
    if request.method == 'POST':
        form = VerificationForm(request.POST)
        if form.is_valid():
            action = request.POST.get('action', 'approve')
            
            # Perform validation checks
            measurements_ok = all([
                4.5 <= checklist.line_pressure <= 5.5,
                11 <= checklist.uv_flow_input_pressure <= 15,
                0.25 <= checklist.test_pressure_vacuum <= 0.3,
                checklist.oring_condition == 'OK',
                checklist.master_verification_lvdt == 'OK',
                checklist.good_bad_master_verification == 'OK',
                checklist.tool_alignment == 'OK'
            ])
            
            # Check subgroup measurements
            subgroups_ok = True
            measurement_issues = []
            for subgroup in checklist.subgroup_entries.all():
                if not (-43 <= subgroup.uv_vacuum_test <= -35):
                    subgroups_ok = False
                    measurement_issues.append(f"Subgroup {subgroup.subgroup_number}: UV vacuum test out of range")
                if not (30 <= subgroup.uv_flow_value <= 40):
                    subgroups_ok = False
                    measurement_issues.append(f"Subgroup {subgroup.subgroup_number}: UV flow value out of range")
            
            if action == 'approve':
                if not measurements_ok or not subgroups_ok:
                    message = "Warning: Some measurements are out of range:\n"
                    if not measurements_ok:
                        message += "- Initial measurements are out of range\n"
                    if measurement_issues:
                        message += "\n".join(measurement_issues)
                    messages.warning(request, message)
                    return render(request, 'main/verify_checklist.html', {
                        'form': form,
                        'checklist': checklist,
                        'verification_type': 'Supervisor',
                        'measurements_ok': measurements_ok,
                        'subgroups_ok': subgroups_ok,
                        'measurement_issues': measurement_issues
                    })
                
                checklist.status = 'supervisor_approved'
                success_message = 'Checklist approved successfully'
            else:
                checklist.status = 'rejected'
                success_message = 'Checklist rejected'
            
            checklist.supervisor_verified_at = timezone.now()
            checklist.supervisor_comments = form.cleaned_data['comments']
            checklist.save()
            
            # Create verification record
            Verification.objects.create(
                checklist=checklist,
                verified_by=request.user,
                verifier_type='supervisor',
                status=checklist.status,
                comments=form.cleaned_data['comments']
            )
            
            messages.success(request, success_message)
            return redirect('supervisor_dashboard')
    else:
        form = VerificationForm()
    
    # Pre-check measurements for the template
    measurements_ok = all([
        4.5 <= checklist.line_pressure <= 5.5,
        11 <= checklist.uv_flow_input_pressure <= 15,
        0.25 <= checklist.test_pressure_vacuum <= 0.3,
        checklist.oring_condition == 'OK',
        checklist.master_verification_lvdt == 'OK',
        checklist.good_bad_master_verification == 'OK',
        checklist.tool_alignment == 'OK'
    ])
    
    measurement_issues = []
    for subgroup in checklist.subgroup_entries.all():
        if not (-43 <= subgroup.uv_vacuum_test <= -35):
            measurement_issues.append(f"Subgroup {subgroup.subgroup_number}: UV vacuum test out of range")
        if not (30 <= subgroup.uv_flow_value <= 40):
            measurement_issues.append(f"Subgroup {subgroup.subgroup_number}: UV flow value out of range")
    
    return render(request, 'main/verify_checklist.html', {
        'form': form,
        'checklist': checklist,
        'verification_type': 'Supervisor',
        'measurements_ok': measurements_ok,
        'measurement_issues': measurement_issues,
        'subgroups': checklist.subgroup_entries.all().order_by('subgroup_number')
    })
    
@login_required
@user_passes_test(lambda u: u.user_type == 'quality_supervisor')
def quality_verify(request, checklist_id):
    checklist = get_object_or_404(ChecklistBase, id=checklist_id)
    
    if request.method == 'POST':
        form = VerificationForm(request.POST)
        if form.is_valid():
            checklist.status = 'quality_approved'
            checklist.quality_verified_at = timezone.now()
            checklist.quality_comments = form.cleaned_data['comments']
            checklist.save()
            
            messages.success(request, 'Quality verification completed')
            return redirect('quality_dashboard')
    else:
        form = VerificationForm()
    
    return render(request, 'main/verify_checklist.html', {
        'form': form,
        'checklist': checklist,
        'verification_type': 'Quality'
    })

# Report Views
@login_required
def daily_report(request):
    date = request.GET.get('date', timezone.now().date())
    if isinstance(date, str):
        date = datetime.strptime(date, '%Y-%m-%d').date()
    
    checklists = ChecklistBase.objects.filter(shift__date=date)
    
    summary = {
        'total': checklists.count(),
        'pending': checklists.filter(status='pending').count(),
        'supervisor_approved': checklists.filter(status='supervisor_approved').count(),
        'quality_approved': checklists.filter(status='quality_approved').count(),
        'rejected': checklists.filter(status='rejected').count(),
    }
    
    return render(request, 'main/reports/daily_report.html', {
        'date': date,
        'checklists': checklists,
        'summary': summary
    })

@login_required
def weekly_report(request):
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=7)
    
    checklists = ChecklistBase.objects.filter(
        shift__date__range=[start_date, end_date]
    ).order_by('shift__date')
    
    daily_stats = {}
    for i in range(7):
        current_date = start_date + timedelta(days=i)
        daily_checklists = checklists.filter(shift__date=current_date)
        
        daily_stats[current_date] = {
            'total': daily_checklists.count(),
            'approved': daily_checklists.filter(status='quality_approved').count(),
            'rejected': daily_checklists.filter(status='rejected').count(),
        }
    
    return render(request, 'main/reports/weekly_report.html', {
        'start_date': start_date,
        'end_date': end_date,
        'daily_stats': daily_stats
    })

@login_required
def monthly_report(request):
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)
    
    checklists = ChecklistBase.objects.filter(
        shift__date__range=[start_date, end_date]
    )
    
    # Calculate statistics by model
    model_stats = {}
    for model in ChecklistBase.MODEL_CHOICES:
        model_checklists = checklists.filter(selected_model=model[0])
        model_stats[model[0]] = {
            'total': model_checklists.count(),
            'approved': model_checklists.filter(status='quality_approved').count(),
            'rejected': model_checklists.filter(status='rejected').count(),
        }
    
    return render(request, 'main/reports/monthly_report.html', {
        'start_date': start_date,
        'end_date': end_date,
        'model_stats': model_stats
    })

# Profile Views
@login_required
def user_profile(request):
    return render(request, 'main/profile/user_profile.html', {
        'user': request.user
    })

@login_required
def edit_profile(request):
    if request.method == 'POST':
        form = UserProfileForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profile updated successfully')
            return redirect('user_profile')
    else:
        form = UserProfileForm(instance=request.user)
    
    return render(request, 'main/profile/edit_profile.html', {
        'form': form
    })

# API Views
@login_required
def validate_checklist(request):
    """API endpoint for validating checklist data with warnings instead of errors"""
    if request.method == 'POST':
        data = request.POST
        warnings = []
        
        try:
            # Validate line pressure
            line_pressure = float(data.get('line_pressure', 0))
            if not (4.5 <= line_pressure <= 5.5):
                warnings.append({
                    'field': 'line_pressure',
                    'message': f'Line pressure value {line_pressure} bar is outside recommended range (4.5 - 5.5 bar)',
                    'value': line_pressure,
                    'recommended_range': {
                        'min': 4.5,
                        'max': 5.5,
                        'unit': 'bar'
                    }
                })
            
            # Validate UV flow test pressure
            uv_pressure = float(data.get('uv_flow_test_pressure', 0))
            if not (11 <= uv_pressure <= 15):
                warnings.append({
                    'field': 'uv_flow_test_pressure',
                    'message': f'UV flow test pressure value {uv_pressure} kPa is outside recommended range (11 - 15 kPa)',
                    'value': uv_pressure,
                    'recommended_range': {
                        'min': 11,
                        'max': 15,
                        'unit': 'kPa'
                    }
                })
            
            # Validate UV vacuum test
            uv_vacuum = float(data.get('uv_vacuum_test', 0))
            if not (-43 <= uv_vacuum <= -35):
                warnings.append({
                    'field': 'uv_vacuum_test',
                    'message': f'UV vacuum test value {uv_vacuum} kPa is outside recommended range (-43 to -35 kPa)',
                    'value': uv_vacuum,
                    'recommended_range': {
                        'min': -43,
                        'max': -35,
                        'unit': 'kPa'
                    }
                })
            
            return JsonResponse({
                'is_valid': True,  # Always valid as we're only showing warnings
                'has_warnings': len(warnings) > 0,
                'warnings': warnings,
                'values': {  # Return validated values for reference
                    'line_pressure': line_pressure,
                    'uv_flow_test_pressure': uv_pressure,
                    'uv_vacuum_test': uv_vacuum
                }
            })
            
        except (TypeError, ValueError) as e:
            return JsonResponse({
                'is_valid': False,
                'errors': ['Please enter valid numerical values'],
                'details': str(e)
            })
    
    return JsonResponse({'error': 'Invalid request method'}, status=400)    
@login_required
def operator_history(request):
    # Get all checklists grouped by date
    checklists = ChecklistBase.objects.filter(
        shift__operator=request.user
    ).annotate(
        subgroup_count=Count('subgroup_entries')
    ).order_by('-created_at')
    
    # Group by status
    pending_checklists = checklists.filter(status='pending')
    verified_checklists = checklists.filter(status__in=['supervisor_approved', 'quality_approved'])
    rejected_checklists = checklists.filter(status='rejected')
    
    context = {
        'title': 'Operator History',
        'pending_checklists': pending_checklists,
        'verified_checklists': verified_checklists,
        'rejected_checklists': rejected_checklists,
        'total_checklists': checklists.count()
    }
    
    return render(request, 'main/history/operator_history.html', context)
@login_required
@user_passes_test(lambda u: u.user_type == 'shift_supervisor')
def supervisor_history(request):
    # Get all verifications by this supervisor
    verifications = SubgroupVerification.objects.filter(
        verified_by=request.user,
        verifier_type='supervisor'
    ).select_related(
        'subgroup__checklist',
        'subgroup__checklist__shift__operator'
    ).order_by('-verified_at')

    # Get subgroups pending verification
    pending_subgroups = SubgroupEntry.objects.filter(
        checklist__shift__shift_supervisor=request.user,
        verification_status='pending'
    ).select_related(
        'checklist',
        'checklist__shift__operator'
    ).order_by('-timestamp')

    return render(request, 'main/history/supervisor_history.html', {
        'pending_verifications': pending_subgroups,
        'verified_entries': verifications,
        'title': 'Supervisor History'
    })
@login_required
@user_passes_test(lambda u: u.user_type == 'quality_supervisor')
def quality_history(request):
    # Get all quality verifications by this supervisor
    verified_entries = SubgroupVerification.objects.filter(
        verified_by=request.user,
        verifier_type='quality'
    ).select_related(
        'subgroup__checklist',
        'subgroup__checklist__shift__operator',
        'subgroup__checklist__shift__shift_supervisor'
    ).order_by('-verified_at')

    # Get subgroups that have supervisor verification but no quality verification
    pending_subgroups = SubgroupEntry.objects.filter(
        verifications__verifier_type='supervisor',
        verifications__status='approved'
    ).exclude(
        verifications__verifier_type='quality'
    ).select_related(
        'checklist',
        'checklist__shift__operator',
        'checklist__shift__shift_supervisor'
    ).prefetch_related(
        'verifications'
    ).order_by('-timestamp')

    context = {
        'pending_verifications': pending_subgroups,
        'verified_entries': verified_entries,
        'title': 'Quality History'
    }

    return render(request, 'main/history/quality_history.html', context)    

@login_required
def user_settings(request):
    if request.method == 'POST':
        # Handle settings update
        notification_settings = request.POST.get('notification_settings', False)
        theme_preference = request.POST.get('theme_preference', 'light')
        
        # Save settings to user profile or settings model
        profile = request.user
        profile.email_notifications = notification_settings == 'on'
        profile.theme_preference = theme_preference
        profile.save()
        
        messages.success(request, 'Settings updated successfully')
        return redirect('user_settings')
    
    return render(request, 'main/profile/user_settings.html', {
        'user': request.user
    })
    
@login_required
def notification_settings(request):
    if request.method == 'POST':
        # Handle notification settings update
        request.user.email_notifications = request.POST.get('email_notifications') == 'on'
        request.user.save()
        messages.success(request, 'Notification settings updated successfully.')
        return redirect('user_settings')
        
    return render(request, 'main/settings/notifications.html', {
        'user': request.user,
        'active_tab': 'notifications'
    })

@login_required
def user_preferences(request):
    if request.method == 'POST':
        # Handle user preferences update
        messages.success(request, 'Preferences updated successfully.')
        return redirect('user_settings')
        
    return render(request, 'main/settings/preferences.html', {
        'user': request.user,
        'active_tab': 'preferences'
    })    
    
    
    
    
@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'])
def verify_subgroup(request, subgroup_id):
    subgroup = get_object_or_404(SubgroupEntry, id=subgroup_id)
    verifier_type = 'supervisor' if request.user.user_type == 'shift_supervisor' else 'quality'
    
    # Check if subgroup can be verified
    if verifier_type == 'quality' and subgroup.verification_status != 'supervisor_verified':
        messages.error(request, 'Subgroup must be verified by supervisor first')
        return redirect('checklist_detail', checklist_id=subgroup.checklist.id)
    
    if request.method == 'POST':
        form = SubgroupVerificationForm(request.POST)
        if form.is_valid():
            verification = form.save(commit=False)
            verification.subgroup = subgroup
            verification.verified_by = request.user
            verification.verifier_type = verifier_type
            
            # Update subgroup status based on verification
            if verification.status == 'rejected':
                subgroup.verification_status = 'rejected'
            elif verifier_type == 'supervisor':
                subgroup.verification_status = 'supervisor_verified'
            else:  # quality supervisor
                subgroup.verification_status = 'quality_verified'
            
            verification.save()
            subgroup.save()
            
            # Check if all subgroups are verified to update checklist status
            update_checklist_status(subgroup.checklist)
            
            messages.success(request, 'Subgroup verified successfully')
            return redirect('checklist_detail', checklist_id=subgroup.checklist.id)
    else:
        form = SubgroupVerificationForm()
    
    context = {
        'form': form,
        'subgroup': subgroup,
        'verifier_type': verifier_type.title(),
        'measurements': {
            'uv_vacuum_test_ok': -43 <= subgroup.uv_vacuum_test <= -35,
            'uv_flow_value_ok': 30 <= subgroup.uv_flow_value <= 40,
            'assembly_ok': subgroup.umbrella_valve_assembly == 'OK',
            'pressing_ok': subgroup.uv_clip_pressing == 'OK'
        }
    }
    
    return render(request, 'main/verify_subgroup.html', context)

def update_checklist_status(checklist):
    """Update checklist status based on subgroup verifications"""
    subgroups = checklist.subgroup_entries.all()
    total_subgroups = subgroups.count()
    
    if total_subgroups == 0:
        return
    
    # Count verifications
    supervisor_verified = subgroups.filter(verification_status='supervisor_verified').count()
    quality_verified = subgroups.filter(verification_status='quality_verified').count()
    rejected = subgroups.filter(verification_status='rejected').count()
    
    # Update checklist status
    if rejected > 0:
        checklist.status = 'rejected'
    elif quality_verified == total_subgroups:
        checklist.status = 'quality_approved'
    elif supervisor_verified == total_subgroups:
        checklist.status = 'supervisor_approved'
    else:
        checklist.status = 'pending'
    
    checklist.save()    